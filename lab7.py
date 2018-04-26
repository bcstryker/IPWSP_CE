import sys
from openpyxl import load_workbook

filename = 'home/student/ipwsp/aws/AWS_SEED.xlsx'
column = 'A'
row = 3

class AWS_VPC():
    def __init__(self):
        try:
            wb = load_workbook(filename = filename, data_only = True)
        except IOError:
            sys.exit('Could not open file')
        else:
            self.wb = wb

    def get_row(self, col, row = 1, start_row = 0, max_row = 4):
        ws = self.wb.active
        sheet = wb['Sheet1']
        if max_row > sheet.max_row:
            print 'Max row referenced out of range.'
            print 'Setting max row to {}'.format(sheet.max_row)
            max_row = sheet.max_row
        test_list = []
        for row in range(start_row, max_row):
            #the following line is confusing and seems unnecessary
            for column in '{}''.format(column):
                cell_name = column + str(row)
                x = sheet[cell_name].value
                test_list.append(x)
            return test_list

def main():
    vpc_console1 = AWS_VPC()
    vpc_list = vpc_console1.get_row('A', start_row = 1, max_row = 7)
    pub_subnet_list = vpc_console1.get_row('B', start_row = 1, max_row = 5)
    priv_subnet_list = vpc_console1.get_row('C', start_row = 1, max_row = 5)
    route_table_list = vpc_console1.get_row('D', start_row = 1, max_row = 5)
    region_list = vpc_console1.get_row('E', start_row = 1, max_row = 5)

if __name__ == '__main__':
    main()
